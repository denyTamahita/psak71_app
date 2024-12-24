from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
import os
import pandas as pd
import numpy as np
from scipy import stats
from datetime import datetime
from dateutil.relativedelta import relativedelta
from dotenv import load_dotenv
from models import db, User, MacroVariable, MacroValue, RegressionResult, MacroMaster, CfgParm, RefPdLtTemp
import re
import statsmodels.api as sm
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill
from io import BytesIO
from flask import send_file

# Load environment variables
load_dotenv()

# Initialize Flask app
app = Flask(__name__)
app.config['SECRET_KEY'] = os.urandom(24)
app.config['SQLALCHEMY_DATABASE_URI'] = (
    f"mssql+pyodbc://{os.getenv('DB_USERNAME')}:{os.getenv('DB_PASSWORD')}@"
    f"{os.getenv('DB_SERVER')}/{os.getenv('DB_NAME')}?"
    "driver=ODBC+Driver+17+for+SQL+Server&"
    "Trusted_Connection=no&"
    "TrustServerCertificate=yes&"
    "Connection+Timeout=30"
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_size': 10,
    'max_overflow': 20,
    'pool_pre_ping': True,
    'pool_recycle': 3600
}

# Initialize extensions
db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Configure file uploads
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), UPLOAD_FOLDER)
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def calculate_time_series(df):
    """Calculate time series components"""
    try:
        # Calculate 4-month moving average
        df['moving_avg_4'] = df['value'].rolling(window=4, center=True).mean()
        
        # Calculate Centered Moving Average (CMA)
        df['cma_4'] = df['moving_avg_4'].rolling(window=2, center=True).mean()
        
        # Calculate Y/CMA ratio
        df['yt_cma'] = df['value'] / df['cma_4']
        
        # Calculate seasonal factors (average Y/CMA for each month)
        seasonal_factors = df.groupby(df.index.month)['yt_cma'].mean()
        
        # Normalize seasonal factors to sum to 12
        if not seasonal_factors.empty and not seasonal_factors.isna().all():
            seasonal_factors = seasonal_factors * (12 / seasonal_factors.sum())
        
        # Add seasonal factors back to the dataframe
        df['seasonal_factor'] = df.index.map(lambda x: seasonal_factors.get(x.month, 1.0))
        
        # Calculate deseasonalized values
        df['deseasonalized'] = df['value'] / df['seasonal_factor'].replace(0, 1)
        
        # Calculate trend using linear regression on non-null values
        X = np.arange(len(df)).reshape(-1, 1)
        y = df['deseasonalized'].values
        mask = ~np.isnan(y)
        
        if np.any(mask):
            model = stats.linregress(X[mask].flatten(), y[mask])
            df['y_pred_deseasonalized'] = model.slope * X.flatten() + model.intercept
            
            # Calculate final predictions (reseasonalized)
            df['y_pred'] = df['y_pred_deseasonalized'] * df['seasonal_factor']
        else:
            df['y_pred_deseasonalized'] = np.nan
            df['y_pred'] = np.nan
        
        # Replace any infinite values with NaN
        df = df.replace([np.inf, -np.inf], np.nan)
        
        return df
        
    except Exception as e:
        print(f"Error in calculate_time_series: {str(e)}")
        # Return dataframe with NaN for calculated columns
        for col in ['moving_avg_4', 'cma_4', 'yt_cma', 'seasonal_factor', 
                   'deseasonalized', 'y_pred_deseasonalized', 'y_pred']:
            df[col] = np.nan
        return df

def process_macro_data(df, macro_var):
    """Process macro variable data with time series analysis"""
    try:
        # Convert index to datetime if not already
        if not isinstance(df.index, pd.DatetimeIndex):
            df.index = pd.to_datetime(df.index)
        
        # Sort by date
        df = df.sort_index()
        
        # Replace NaN values with None before processing
        df = df.replace({np.nan: None})
        
        # Skip rows with None values
        valid_data = df[df['value'].notna()].copy()
        
        if len(valid_data) > 0:
            # Calculate time series components only for valid data
            valid_data = calculate_time_series(valid_data)
            
            # Create a mapping of dates to calculated values
            calculated_values = valid_data.to_dict('index')
            
            # Store values in database
            for idx, row in df.iterrows():
                # Get calculated values if they exist
                calc_row = calculated_values.get(idx, {})
                
                # Only insert if we have a valid value
                if row['value'] is not None:
                    try:
                        value = MacroValue(
                            macro_variable_id=macro_var.id,
                            date=idx.date(),
                            value=float(row['value']),
                            moving_avg_4=float(calc_row.get('moving_avg_4')) if calc_row.get('moving_avg_4') is not None else None,
                            cma_4=float(calc_row.get('cma_4')) if calc_row.get('cma_4') is not None else None,
                            yt_cma=float(calc_row.get('yt_cma')) if calc_row.get('yt_cma') is not None else None,
                            seasonal_factor=float(calc_row.get('seasonal_factor')) if calc_row.get('seasonal_factor') is not None else None,
                            deseasonalized=float(calc_row.get('deseasonalized')) if calc_row.get('deseasonalized') is not None else None,
                            y_pred_deseasonalized=float(calc_row.get('y_pred_deseasonalized')) if calc_row.get('y_pred_deseasonalized') is not None else None,
                            y_pred=float(calc_row.get('y_pred')) if calc_row.get('y_pred') is not None else None,
                            is_predicted=False
                        )
                        db.session.add(value)
                    except (ValueError, TypeError) as e:
                        print(f"Error processing row for date {idx}: {str(e)}")
                        continue
        
        db.session.commit()
            
    except Exception as e:
        db.session.rollback()
        raise Exception(f"Error processing data: {str(e)}")

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('macro_variables'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            login_user(user)
            return redirect(url_for('macro_variables'))
        
        flash('Invalid username or password')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/upload_macro', methods=['GET', 'POST'])
@login_required
def upload_macro():
    # Get unique periods from macro_master
    periods = db.session.query(MacroMaster.periode).distinct().order_by(MacroMaster.periode.desc()).all()
    selected_period = request.args.get('periode', periods[0][0] if periods else None)
    
    # Get macro_master data for selected period
    macro_master_data = None
    if selected_period:
        macro_master_data = MacroMaster.query.filter_by(periode=selected_period)\
            .order_by(MacroMaster.macro_variable_id).all()
    
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file part', 'error')
            return redirect(request.url)
            
        file = request.files['file']
        if file.filename == '':
            flash('Pilih file terlebih dahulu', 'error')
            return redirect(request.url)
        
        # Get period
        period = request.form.get('periode')
        
        if not period:
            flash('Pilih periode', 'error')
            return redirect(request.url)
            
        if file and allowed_file(file.filename):
            try:
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                
                # Read the file
                try:
                    if filename.endswith('.csv'):
                        df = pd.read_csv(filepath)
                    else:
                        df = pd.read_excel(filepath)
                    
                    print("File contents before processing:")
                    print("Columns:", df.columns.tolist())
                    print("Shape:", df.shape)
                    print("First few rows:")
                    print(df.head())
                    print("\nAll rows:")
                    print(df)
                    
                    # Drop header row if it exists
                    if df.shape[0] > 0:
                        # Check if first row contains headers
                        if isinstance(df.iloc[0,0], str) and not df.iloc[0,0].isdigit():
                            df = df.iloc[1:]
                            df = df.reset_index(drop=True)
                    
                    print("\nFile contents after dropping header:")
                    print("Shape:", df.shape)
                    print("First few rows:")
                    print(df.head())
                    print("\nAll rows after header removal:")
                    print(df)
                    
                    if df.empty:
                        raise ValueError("File tidak memiliki data setelah header dihapus")
                    
                    # Get macro_master data for reference
                    macro_master = MacroMaster.query.filter_by(
                        periode=period
                    ).order_by(MacroMaster.macro_variable_id).all()
                    
                    print("\nMacro master data:")
                    for m in macro_master:
                        print(f"ID: {m.macro_variable_id}, Name: {m.macro_variable_name}")
                    
                    if len(df.columns) != len(macro_master) + 2:
                        raise ValueError(f"File harus memiliki {len(macro_master) + 2} kolom: PERIODE, DATE_REGRESI, dan nilai untuk setiap variabel. File memiliki {len(df.columns)} kolom")
                    
                    # Drop any rows with all NaN values
                    df = df.dropna(how='all')
                    
                    # Process each row
                    successful_rows = 0
                    total_rows = len(df)
                    
                    for idx, row in df.iterrows():
                        try:
                            # Skip row if first two columns are NaN
                            if pd.isna(row[0]) or pd.isna(row[1]):
                                print(f"Skipping row {idx+1} - Missing periode or date_regresi")
                                continue
                                
                            print(f"\nProcessing row {idx+1} of {total_rows}:")
                            print("Row data:", row.tolist())
                            
                            try:
                                upload_period = str(int(float(str(row[0]).strip())))
                                date_regresi_str = str(int(float(str(row[1]).strip())))
                            except (ValueError, TypeError) as e:
                                print(f"Error converting periode/date on row {idx+1}: {str(e)}")
                                continue
                            
                            print(f"Periode: {upload_period}")
                            print(f"Date Regresi String: {date_regresi_str}")
                            
                            # Convert date_regresi to numeric format
                            try:
                                if len(date_regresi_str) == 8:  # YYYYMMDD
                                    date_obj = datetime.strptime(date_regresi_str, '%Y%m%d')
                                    date_regresi = int(date_obj.strftime('%Y%m'))
                                elif len(date_regresi_str) == 6:  # YYYYMM
                                    date_regresi = int(date_regresi_str)
                                else:
                                    raise ValueError(f"Invalid date format: {date_regresi_str}")
                                print(f"Converted Date Regresi: {date_regresi}")
                            except ValueError as e:
                                print(f"Date conversion error on row {idx+1}: {str(e)}")
                                continue
                            
                            # Validate periode format
                            if not re.match(r'^\d{6}$', upload_period):
                                print(f"Invalid periode format on row {idx+1}")
                                continue
                            
                            # Create values for each macro variable
                            row_valid = True
                            for i, macro in enumerate(macro_master):
                                try:
                                    if pd.isna(row[i + 2]):
                                        print(f"Missing value for {macro.macro_variable_name} on row {idx+1}")
                                        row_valid = False
                                        break
                                        
                                    value = float(row[i + 2])
                                    print(f"Variable {macro.macro_variable_name}: {value}")
                                    
                                    # Delete existing value with same primary key if exists
                                    MacroValue.query.filter_by(
                                        periode=upload_period,
                                        macro_variable_id=macro.macro_variable_id,
                                        date_regresi=date_regresi
                                    ).delete()
                                    
                                    macro_value = MacroValue(
                                        periode=upload_period,
                                        macro_variable_id=macro.macro_variable_id,
                                        macro_variable_name=macro.macro_variable_name,
                                        date_regresi=date_regresi,
                                        value=value
                                    )
                                    db.session.add(macro_value)
                                    db.session.flush()
                                    
                                except (ValueError, TypeError) as e:
                                    print(f"Error on column {i+3} (variable {macro.macro_variable_name}): {str(e)}")
                                    row_valid = False
                                    break
                            
                            if row_valid:
                                successful_rows += 1
                                db.session.commit()  # Commit each successful row
                            else:
                                db.session.rollback()
                            
                        except Exception as e:
                            print(f"Error processing row {idx+1}: {str(e)}")
                            db.session.rollback()
                            continue
                    
                    if successful_rows == 0:
                        flash('Tidak ada data yang berhasil diupload', 'error')
                    else:
                        flash(f'Berhasil mengupload {successful_rows} dari {total_rows} baris data', 'success')
                    return redirect(url_for('macro_variables'))
                    
                except pd.errors.EmptyDataError:
                    flash('File yang diupload kosong', 'error')
                    return redirect(request.url)
                    
                except pd.errors.ParserError as e:
                    flash(f'Error membaca file. Periksa format file. Error: {str(e)}', 'error')
                    return redirect(request.url)
                except ValueError as e:
                    flash(f'Error: {str(e)}', 'error')
                    return redirect(request.url)
                except Exception as e:
                    print(f"Unexpected error: {str(e)}")
                    print(f"Error type: {type(e)}")
                    import traceback
                    print("Traceback:")
                    print(traceback.format_exc())
                    db.session.rollback()
                    flash(f'Error tidak terduga: {str(e)}', 'error')
                    return redirect(request.url)
                    
            except Exception as e:
                db.session.rollback()
                flash(f'Error memproses file: {str(e)}', 'error')
                return redirect(request.url)
            
            finally:
                # Clean up uploaded file
                if os.path.exists(filepath):
                    os.remove(filepath)
    
    return render_template('upload_macro_data.html', 
                         periods=[p[0] for p in periods],
                         selected_period=selected_period,
                         macro_master=macro_master_data)

@app.route('/upload_macro_data', methods=['GET', 'POST'])
@login_required
def upload_macro_data():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file uploaded', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)
        
        if file and allowed_file(file.filename):
            try:
                # Read Excel file
                df = pd.read_excel(file)
                
                # Get periode from form
                periode = request.form.get('periode')
                if not periode:
                    flash('Periode is required', 'error')
                    return redirect(request.url)
                
                # Convert periode to numeric format
                periode = int(periode.replace('-', ''))
                
                # Process each row
                for _, row in df.iterrows():
                    # Check if macro variable exists
                    macro_var = MacroMaster.query.filter_by(
                        macro_variable_id=row['macro_variable_id'],
                        periode=periode
                    ).first()
                    
                    if not macro_var:
                        # Create new macro variable
                        macro_var = MacroMaster(
                            macro_variable_id=row['macro_variable_id'],
                            macro_variable_name=row['macro_variable_name'],
                            periode=periode
                        )
                        db.session.add(macro_var)
                    
                    # Process date_regresi values
                    date_str = str(row['date_regresi'])
                    # Convert date to YYYYMM format
                    if '-' in date_str:
                        # If date is in YYYY-MM-DD format
                        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                        date_numeric = int(date_obj.strftime('%Y%m'))
                    else:
                        # If date is already in YYYYMM format
                        date_numeric = int(date_str)
                    
                    # Create or update macro value
                    macro_value = MacroValue.query.filter_by(
                        macro_variable_id=row['macro_variable_id'],
                        periode=periode,
                        date_regresi=date_numeric
                    ).first()
                    
                    if not macro_value:
                        macro_value = MacroValue(
                            macro_variable_id=row['macro_variable_id'],
                            macro_variable_name=row['macro_variable_name'],
                            periode=periode,
                            date_regresi=date_numeric,
                            value=row['value']
                        )
                        db.session.add(macro_value)
                    else:
                        macro_value.value = row['value']
                
                db.session.commit()
                flash('Data uploaded successfully', 'success')
                return redirect(url_for('macro_variables'))  # Changed from macro_data to macro_variables
                
            except Exception as e:
                db.session.rollback()
                flash(f'Error uploading data: {str(e)}', 'error')
                return redirect(request.url)
    
    return render_template('upload_macro_data.html')

@app.route('/macro_variables')
@login_required
def macro_variables():
    current_period = request.args.get('periode', None)
    if not current_period:
        # Get latest period
        latest = db.session.query(MacroValue.periode)\
            .order_by(MacroValue.periode.desc())\
            .first()
        current_period = latest[0] if latest else None

    # Get variables for the period
    variables = []
    if current_period:
        variables = db.session.query(
            MacroValue.macro_variable_id,
            MacroValue.macro_variable_name,
            MacroValue.value,
            MacroValue.date_regresi
        ).filter_by(periode=current_period).all()

    return render_template('macro_variables.html', 
                         variables=variables,
                         current_period=current_period)

@app.route('/macro_variable_detail/<periode>/<int:macro_variable_id>')
@login_required
def macro_variable_detail(periode, macro_variable_id):
    try:
        # Get macro master data
        macro = MacroMaster.query.filter_by(
            periode=periode,
            macro_variable_id=macro_variable_id
        ).first_or_404()
        
        # Get all values for this variable
        values = MacroValue.query.filter_by(
            periode=periode,
            macro_variable_id=macro_variable_id
        ).order_by(MacroValue.date_regresi).all()
        
        # Calculate statistics if there are values
        stats = {}
        if values:
            values_list = [v.value for v in values]
            stats['min'] = min(values_list)
            stats['max'] = max(values_list)
            stats['avg'] = sum(values_list) / len(values_list)
            stats['count'] = len(values_list)
            stats['latest'] = values[-1].value
            stats['latest_date'] = values[-1].date_regresi
        
        return render_template(
            'macro_variable_detail.html',
            macro=macro,
            values=values,
            stats=stats,
            periode=periode
        )
        
    except Exception as e:
        flash(f'Error loading macro variable details: {str(e)}', 'error')
        return redirect(url_for('macro_variables'))

@app.route('/regression/<periode>/<int:macro_variable_id>', methods=['GET', 'POST'])
@login_required
def regression(periode, macro_variable_id):
    # Get macro master data
    macro = MacroMaster.query.filter_by(
        periode=periode,
        macro_variable_id=macro_variable_id
    ).first_or_404()
    
    # Get all values for this variable
    values = MacroValue.query.filter_by(
        periode=periode,
        macro_variable_id=macro_variable_id
    ).order_by(MacroValue.date_regresi).all()
    
    if not values:
        flash('Tidak ada data untuk analisis regresi', 'error')
        return redirect(url_for('macro_variables'))
    
    if request.method == 'POST':
        try:
            # Get selected variables for regression
            dependent_var = request.form.get('dependent_var')
            independent_vars = request.form.getlist('independent_vars')
            
            if not dependent_var or not independent_vars:
                flash('Pilih variabel dependen dan independen', 'error')
                return redirect(url_for('regression', periode=periode, macro_variable_id=macro_variable_id))
            
            # Get data for regression
            regression_data = {}
            
            # Get dependent variable data
            dep_values = MacroValue.query.filter_by(
                periode=periode,
                macro_variable_id=int(dependent_var)
            ).order_by(MacroValue.date_regresi).all()
            
            if not dep_values:
                flash('Data variabel dependen tidak ditemukan', 'error')
                return redirect(url_for('regression', periode=periode, macro_variable_id=macro_variable_id))
            
            regression_data['y'] = [v.value for v in dep_values]
            regression_data['dates'] = [v.date_regresi.strftime('%Y-%m-%d') for v in dep_values]
            
            # Get independent variables data
            X_data = []
            var_names = []
            for var_id in independent_vars:
                ind_values = MacroValue.query.filter_by(
                    periode=periode,
                    macro_variable_id=int(var_id)
                ).order_by(MacroValue.date_regresi).all()
                
                if not ind_values:
                    flash(f'Data variabel independen {var_id} tidak ditemukan', 'error')
                    return redirect(url_for('regression', periode=periode, macro_variable_id=macro_variable_id))
                
                X_data.append([v.value for v in ind_values])
                var_names.append(ind_values[0].macro_variable_name)
            
            # Convert to numpy arrays
            X = np.array(X_data).T
            y = np.array(regression_data['y'])
            
            # Perform multiple linear regression
            model = stats.linregress(X, y) if len(X_data) == 1 else stats.LinearRegression().fit(X, y)
            
            # Calculate predictions
            y_pred = model.predict(X) if len(X_data) > 1 else model.slope * X.flatten() + model.intercept
            
            # Calculate R-squared
            r_squared = model.rvalue ** 2 if len(X_data) == 1 else model.score(X, y)
            
            # Prepare results for template
            results = {
                'r_squared': r_squared,
                'coefficients': model.slope if len(X_data) == 1 else model.coef_,
                'intercept': model.intercept,
                'variable_names': var_names,
                'actual_values': y.tolist(),
                'predicted_values': y_pred.tolist(),
                'dates': [d.strftime('%Y-%m-%d') for d in regression_data['dates']]
            }
            
            return render_template(
                'regression_results.html',
                macro=macro,
                results=results,
                periode=periode
            )
            
        except Exception as e:
            flash(f'Error melakukan regresi: {str(e)}', 'error')
            return redirect(url_for('regression', periode=periode, macro_variable_id=macro_variable_id))
    
    # Get all available variables for this period
    available_vars = MacroMaster.query.filter_by(periode=periode).all()
    
    return render_template(
        'regression.html',
        macro=macro,
        values=values,
        available_vars=available_vars,
        periode=periode
    )

def get_pd_template():
    try:
        template_path = os.path.join(app.root_path, 'templates', 'pd_template.csv')
        df = pd.read_csv(template_path)
        return df['value'].tolist()
    except Exception as e:
        print(f"Error reading PD template: {str(e)}")
        return []

@app.route('/get_macro_values/<periode>/<int:macro_id>')
@login_required
def get_macro_values(periode, macro_id):
    try:
        values = MacroValue.query.filter_by(
            periode=periode,
            macro_variable_id=macro_id
        ).order_by(MacroValue.date_regresi).all()
        
        return jsonify({
            'values': [float(v.value) for v in values],
            'dates': [v.date_regresi.strftime('%Y-%m-%d') for v in values]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/regression_test', methods=['GET', 'POST'])
@login_required
def regression_test():
    # Get all periods
    periods = db.session.query(MacroValue.periode).distinct().order_by(MacroValue.periode.desc()).all()
    periods = [p[0] for p in periods]
    
    # Get current period
    current_period = request.args.get('periode', periods[0] if periods else None)
    
    # Get all variables for the period
    variables = db.session.query(
        MacroMaster.macro_variable_id,
        MacroMaster.macro_variable_name
    ).filter_by(periode=current_period).all()
    
    if request.method == 'POST':
        periode = request.form.get('periode')
        y_variable = request.form.get('y_variable')
        x_variables = request.form.getlist('x_variables')
        
        # Get PD values for selected X variables
        pd_values = {}
        for x_var in x_variables:
            pd_value = request.form.get(f'pd_{x_var}')
            if pd_value:
                pd_values[int(x_var)] = float(pd_value)
        
        if y_variable and x_variables:
            # Get data for regression
            y_data = MacroValue.query.filter_by(
                periode=periode,
                macro_variable_id=int(y_variable)
            ).order_by(MacroValue.date_regresi).all()
            
            # Create X and y arrays for regression
            dates = [value.date_regresi for value in y_data]
            y = np.array([value.value for value in y_data])
            X = []
            var_names = []
            
            for x_var in x_variables:
                x_values = MacroValue.query.filter_by(
                    periode=periode,
                    macro_variable_id=int(x_var)
                ).order_by(MacroValue.date_regresi).all()
                
                if len(x_values) == len(y_data):
                    X.append([value.value for value in x_values])
                    var_name = MacroMaster.query.filter_by(
                        periode=periode,
                        macro_variable_id=int(x_var)
                    ).first().macro_variable_name
                    var_names.append(var_name)
            
            X = np.array(X).T
            
            # Perform regression
            model = sm.OLS(y, sm.add_constant(X))
            results = model.fit()
            
            # Prepare results dictionary
            regression_results = {
                'r': np.sqrt(results.rsquared),
                'r_squared': results.rsquared,
                'adj_r_squared': results.rsquared_adj,
                'std_error': np.sqrt(results.mse_resid),
                
                'anova': {
                    'regression_ss': results.ess,
                    'regression_df': results.df_model,
                    'regression_ms': results.ess / results.df_model,
                    'residual_ss': results.ssr,
                    'residual_df': results.df_resid,
                    'residual_ms': results.ssr / results.df_resid,
                    'total_ss': results.ess + results.ssr,
                    'total_df': results.df_model + results.df_resid,
                    'f_stat': results.fvalue,
                    'f_pvalue': results.f_pvalue
                },
                
                'coefficients': {
                    'intercept': {
                        'coef': results.params[0],
                        'std_err': results.bse[0],
                        't_stat': results.tvalues[0],
                        'p_value': results.pvalues[0],
                        'lower_ci': results.conf_int()[0][0],
                        'upper_ci': results.conf_int()[1][0]
                    },
                    'variables': []
                },
                
                'data': {
                    'dates': [d.strftime('%Y-%m-%d') for d in dates],
                    'y_actual': y.tolist(),
                    'y_pred': results.predict().tolist()
                }
            }
            
            # Add coefficients for each variable
            for i, var_name in enumerate(var_names):
                var_coef = {
                    'name': var_name,
                    'coef': results.params[i+1],
                    'std_err': results.bse[i+1],
                    't_stat': results.tvalues[i+1],
                    'p_value': results.pvalues[i+1],
                    'lower_ci': results.conf_int()[0][i+1],
                    'upper_ci': results.conf_int()[1][i+1]
                }
                regression_results['coefficients']['variables'].append(var_coef)
            
            # Calculate prediction with PD values
            pd_prediction = {
                'variables': [],
                'total': results.params[0]  # Start with intercept
            }
            
            for i, x_var in enumerate(x_variables):
                if int(x_var) in pd_values:
                    impact = results.params[i+1] * pd_values[int(x_var)]
                    pd_prediction['variables'].append({
                        'name': var_names[i],
                        'pd_value': pd_values[int(x_var)],
                        'coefficient': results.params[i+1],
                        'impact': impact
                    })
                    pd_prediction['total'] += impact
            
            regression_results['pd_prediction'] = pd_prediction
            
            return render_template('regression_test.html',
                                 periods=periods,
                                 periode=periode,
                                 variables=variables,
                                 y_variable=int(y_variable),
                                 x_variables=[int(x) for x in x_variables],
                                 pd_values=pd_values,
                                 results=regression_results)
    
    return render_template('regression_test.html',
                         periods=periods,
                         periode=current_period,
                         variables=variables,
                         pd_values={})

@app.route('/regression_data', methods=['GET', 'POST'])
@login_required
def regression_data():
    # Get all periods
    periods = db.session.query(MacroValue.periode).distinct().order_by(MacroValue.periode.desc()).all()
    periods = [p[0] for p in periods]
    
    # Get current period
    current_period = request.args.get('periode', periods[0] if periods else None)
    if request.method == 'POST':
        current_period = request.form.get('periode', current_period)
    
    # Get segmentasi list from cfg_parm
    segmentasi_list = CfgParm.query.filter_by(parmgrp='101').order_by(CfgParm.parmid).all()
    selected_segmentasi = request.form.get('segmentasi', '01')
    print(f"Selected Segmentasi: {selected_segmentasi}")
    print(f"Request Form Data: {request.form}")
    
    # Get macro variables for the period
    variables = db.session.query(
        MacroMaster.macro_variable_id,
        MacroMaster.macro_variable_name
    ).filter_by(periode=current_period).all()
    
    if request.method == 'POST':
        x_variables = request.form.getlist('x_variables')
        
        if x_variables:
            # Get data for each selected X variable
            x_data = {}
            selected_vars = []
            
            for x_var in x_variables:
                # Get variable data
                var_data = db.session.query(
                    MacroValue.date_regresi,
                    MacroValue.value
                ).filter_by(
                    periode=current_period,
                    macro_variable_id=int(x_var)
                ).order_by(MacroValue.date_regresi).all()
                
                # Get variable name
                var_name = next((var[1] for var in variables if var[0] == int(x_var)), '')
                
                x_data[x_var] = var_data
                selected_vars.append({
                    'id': int(x_var),
                    'name': var_name
                })
            
            # Get date range from first variable
            if x_data:
                first_var = list(x_data.keys())[0]
                start_date = x_data[first_var][0][0]  # Already in numeric format
                end_date = x_data[first_var][-1][0]   # Already in numeric format
                
                # Get PD data
                pd_data = db.session.query(
                    RefPdLtTemp.periode,
                    RefPdLtTemp.pd_pct
                ).filter(
                    RefPdLtTemp.periode.between(start_date, end_date),
                    RefPdLtTemp.version == 0,
                    RefPdLtTemp.rating == 1,
                    RefPdLtTemp.prodid == selected_segmentasi,
                    RefPdLtTemp.pd_seq == '1'
                ).order_by(RefPdLtTemp.periode).all()
                
                # Create pd_dict for easy lookup
                pd_dict = {pd.periode: pd.pd_pct for pd in pd_data}
                
                # Combine data
                data = []
                for i in range(len(x_data[first_var])):
                    date_numeric = x_data[first_var][i][0]
                    # Format date as YYYYMM
                    display_date = str(int(date_numeric))  # This will show as YYYYMM
                    
                    row = {
                        'date': display_date,
                        'ypd': pd_dict.get(date_numeric, 0)
                    }
                    
                    # Add value for each X variable
                    for var_id in x_data:
                        if i < len(x_data[var_id]):
                            row[f'x_{var_id}'] = x_data[var_id][i][1]
                    
                    data.append(row)
            
            return render_template('regression_vars.html',
                                 periods=periods,
                                 selected_period=current_period,
                                 segmentasi_list=segmentasi_list,
                                 selected_segmentasi=selected_segmentasi,
                                 variables=variables,
                                 selected_x=list(map(int, x_variables)),
                                 selected_vars=selected_vars,
                                 data=data)
    
    return render_template('regression_vars.html',
                         periods=periods,
                         selected_period=current_period,
                         segmentasi_list=segmentasi_list,
                         selected_segmentasi=selected_segmentasi,
                         variables=variables,
                         selected_x=[],
                         data=None)

@app.route('/run_regression', methods=['POST'])
@login_required
def run_regression_analysis():
    try:
        data = request.get_json()
        
        # Extract data
        x_variables = data['x_variables']
        raw_data = data['data']
        
        # Prepare data for regression
        X = []
        y = []
        
        for row in raw_data:
            x_row = []
            for var_id in x_variables:
                x_row.append(row[f'x_{var_id}'])
            if all(x is not None for x in x_row) and row['ypd'] is not None:
                X.append(x_row)
                y.append(row['ypd'])
        
        X = sm.add_constant(np.array(X))
        y = np.array(y)
        
        # Perform regression
        model = sm.OLS(y, X)
        results = model.fit()
        
        # Calculate additional statistics
        n = len(y)  # number of observations
        k = len(x_variables)  # number of independent variables
        
        # Degrees of freedom
        df_reg = k
        df_resid = n - k - 1
        df_total = n - 1
        
        # Sum of squares
        ss_total = np.sum((y - np.mean(y))**2)
        ss_resid = np.sum(results.resid**2)
        ss_reg = ss_total - ss_resid
        
        # Mean squares
        ms_reg = ss_reg / df_reg
        ms_resid = ss_resid / df_resid
        
        # Get variable names
        var_names = []
        for var_id in x_variables:
            var = MacroMaster.query.filter_by(
                periode=data['periode'],
                macro_variable_id=var_id
            ).first()
            if var:
                var_names.append(var.macro_variable_name)
        
        # Format results
        regression_results = {
            'multiple_r': float(np.sqrt(results.rsquared)),
            'r_squared': float(results.rsquared),
            'adjusted_r_squared': float(results.rsquared_adj),
            'standard_error': float(np.sqrt(ms_resid)),
            'observations': n,
            
            # ANOVA statistics
            'regression_df': int(df_reg),
            'regression_ss': float(ss_reg),
            'regression_ms': float(ms_reg),
            'residual_df': int(df_resid),
            'residual_ss': float(ss_resid),
            'residual_ms': float(ms_resid),
            'total_df': int(df_total),
            'total_ss': float(ss_total),
            'f_statistic': float(results.fvalue),
            'f_pvalue': float(results.f_pvalue),
            
            'variables': []
        }
        
        # Add variable statistics
        # First row for intercept
        conf_int = results.conf_int()
        regression_results['variables'].append({
            'name': 'Intercept',
            'coefficient': float(results.params[0]),
            'std_err': float(results.bse[0]),
            't_stat': float(results.tvalues[0]),
            'p_value': float(results.pvalues[0]),
            'lower_95': float(conf_int[0][0]),
            'upper_95': float(conf_int[0][1]),
            'lower_95_0': float(conf_int[0][0]),
            'upper_95_0': float(conf_int[0][1])
        })
        
        # Then for each variable
        for i, var_name in enumerate(var_names):
            regression_results['variables'].append({
                'name': var_name,
                'coefficient': float(results.params[i+1]),
                'std_err': float(results.bse[i+1]),
                't_stat': float(results.tvalues[i+1]),
                'p_value': float(results.pvalues[i+1]),
                'lower_95': float(conf_int[i+1][0]),
                'upper_95': float(conf_int[i+1][1]),
                'lower_95_0': float(conf_int[i+1][0]),
                'upper_95_0': float(conf_int[i+1][1])
            })
        
        return jsonify(regression_results)
        
    except Exception as e:
        print(f"Error in regression: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/export_regression', methods=['POST'])
@login_required
def export_regression():
    try:
        results = request.get_json()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Regression Results"
        
        # Style for headers
        header_style = NamedStyle(name='header_style')
        header_style.font = Font(bold=True)
        header_style.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        
        # Add Regression Statistics
        ws.append(['SUMMARY OUTPUT'])
        ws.append([])
        ws.append(['Regression Statistics'])
        ws.append(['Multiple R', results['multiple_r']])
        ws.append(['R Square', results['r_squared']])
        ws.append(['Adjusted R Square', results['adjusted_r_squared']])
        ws.append(['Standard Error', results['standard_error']])
        ws.append(['Observations', results['observations']])
        ws.append([])
        
        # Add ANOVA table
        ws.append(['ANOVA'])
        headers = ['', 'df', 'SS', 'MS', 'F', 'Significance F']
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.style = header_style
            
        ws.append(['Regression', 
                  results['regression_df'],
                  results['regression_ss'],
                  results['regression_ms'],
                  results['f_statistic'],
                  results['f_pvalue']])
        ws.append(['Residual',
                  results['residual_df'],
                  results['residual_ss'],
                  results['residual_ms'],
                  '',
                  ''])
        ws.append(['Total',
                  results['total_df'],
                  results['total_ss'],
                  '',
                  '',
                  ''])
        ws.append([])
        
        # Add Coefficients table
        coeff_headers = ['', 'Coefficients', 'Standard Error', 't Stat', 'P-value',
                        'Lower 95%', 'Upper 95%', 'Lower 95.0%', 'Upper 95.0%']
        ws.append(coeff_headers)
        for cell in ws[ws.max_row]:
            cell.style = header_style
            
        for var in results['variables']:
            ws.append([var['name'],
                      var['coefficient'],
                      var['std_err'],
                      var['t_stat'],
                      var['p_value'],
                      var['lower_95'],
                      var['upper_95'],
                      var['lower_95_0'],
                      var['upper_95_0']])
        
        # Format numbers
        for row in ws.iter_rows(min_row=4, max_row=8):  # Regression Statistics
            if isinstance(row[1].value, (int, float)):
                row[1].number_format = '0.00000000'
                
        for row in ws.iter_rows(min_row=12, max_row=14):  # ANOVA table
            for cell in row[1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00000000'
                    
        for row in ws.iter_rows(min_row=17):  # Coefficients table
            for cell in row[1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00000000'
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save to memory
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='regression_results.xlsx'
        )
        
    except Exception as e:
        print(f"Error exporting to Excel: {str(e)}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)
